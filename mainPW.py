# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'mainwindow.ui'
#
# Created by: PyQt5 UI code generator 5.11.3
#
# WARNING! All changes made in this file will be lost!
import os
import sys
import time

import pyautogui
from PIL import Image as PImage
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QPixmap, QTransform, QIcon
from PyQt5.QtWidgets import QFileDialog, QApplication, QMainWindow, QScrollArea, QGraphicsView
from openpyxl import Workbook

import pdf2image
from baiduocr import BaiDuOcr

import mtp2autocad



class Ui_MainWindow(QMainWindow):

    def __init__(self,parent=None):
        super(Ui_MainWindow,self).__init__(parent)

        self.inx=0
        self.inxle=0
        self.pageNum=0
        self.gw = pyautogui.size()[0]
        self.gh = pyautogui.size()[1]
        self.posx1 = 0
        self.posy1 = 0
        self.posx2 = 0
        self.posy2 = 0
        self.sx1=0
        self.sx2=0
        self.sy1=0
        self.sy2=0
        self.dist=0
        self.bhdu=True
        self.ffname=''


        self.relWidth = 0
        self.relHeight = 0
        self.count = 0
        self.listdata = []
        self.aabb=False
        self.fname=""
        self.finame=''
        self.nsizew=0
        self.nsizeh=0
        self.npixmap=QPixmap()
        self.openmethod=""
        self.img_name_convert_saved="p2imtcs.png"
        self.img_name_rotate_left="p2imtrl.png"
        self.img_name_rotate_right='p2imtrr.png'
        self.img_name_zoomin_saved='p2imtzi.png'
        self.img_name_zoomout_saved='p2imtzo.png'
        self.img_name_orc_invoke='p2imtorci.jpg'
        self.ls_content=[]
        self.bdocr = BaiDuOcr()
        self.mtdwg = None



        self.setupUi(self)



    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        self.mw=self.gw-200
        self.mh=self.gh-200
        self.mposx=(self.gw-self.mw)/2
        self.mposy=(self.gh-self.mh-130)/2
        MainWindow.setGeometry(self.mposx,self.mposy,self.mw,self.mh)
        # MainWindow.resize(self.gw-200, self.gh-50)
        self.centralWidget = QtWidgets.QWidget(MainWindow)
        self.centralWidget.setObjectName("centralWidget")
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout(self.centralWidget)
        self.horizontalLayout_3.setContentsMargins(10, 11, 10, 11)
        self.horizontalLayout_3.setSpacing(6)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setSpacing(6)
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setSpacing(6)

        self.horizontalLayout.setObjectName("horizontalLayout")


        self.btn_loadpdf = QtWidgets.QPushButton(self.centralWidget)
        self.btn_loadpdf.setObjectName("btn_loadpdf")
        self.horizontalLayout.addWidget(self.btn_loadpdf)

        self.btn_loadimg=QtWidgets.QPushButton(self.centralWidget)
        self.btn_loadimg.setObjectName("btn_loadimg")
        self.horizontalLayout.addWidget(self.btn_loadimg)

        self.btn_convert = QtWidgets.QPushButton(self.centralWidget)
        self.btn_convert.setObjectName("btn_convert")
        self.horizontalLayout.addWidget(self.btn_convert)

        self.btn_nextpage = QtWidgets.QPushButton(self.centralWidget)
        self.btn_nextpage.setObjectName("btn_nextpage")
        self.horizontalLayout.addWidget(self.btn_nextpage)

        self.btn_prepage=QtWidgets.QPushButton(self.centralWidget)
        self.btn_prepage.setObjectName("btn_prepage")
        self.horizontalLayout.addWidget(self.btn_prepage)

        self.btn_zoomin=QtWidgets.QPushButton(self.centralWidget)
        self.btn_zoomin.setObjectName('btn_zoomin')
        self.horizontalLayout.addWidget(self.btn_zoomin)

        self.btn_zoomout = QtWidgets.QPushButton(self.centralWidget)
        self.btn_zoomout.setObjectName('btn_zoomout')
        self.horizontalLayout.addWidget(self.btn_zoomout)

        self.btn_zoomn=QtWidgets.QPushButton(self.centralWidget)
        self.btn_zoomn.setObjectName('btn_zoomn')
        self.horizontalLayout.addWidget(self.btn_zoomn)

        self.btn_rotationLeft = QtWidgets.QPushButton(self.centralWidget)
        self.btn_rotationLeft .setObjectName('btn_rotationLeft ')
        self.horizontalLayout.addWidget(self.btn_rotationLeft )

        self.btn_rotationright = QtWidgets.QPushButton(self.centralWidget)
        self.btn_rotationright.setObjectName('btn_rotationright')
        self.horizontalLayout.addWidget(self.btn_rotationright)

        self.le_data=QtWidgets.QLineEdit(self.centralWidget)
        self.le_data.setObjectName('le_data')
        self.horizontalLayout.addWidget(self.le_data)


        self.btn_zoomout.setEnabled(True)
        self.btn_zoomin.setEnabled(True)
        self.btn_rotationLeft.setEnabled(True)
        self.btn_rotationright.setEnabled(True)


        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setSpacing(6)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")

        self.scrollArea = QtWidgets.QScrollArea(self.centralWidget)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, self.gw, self.gh))



        # self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, self.mw, self.gh-5))
        self.scrollAreaWidgetContents.setMinimumSize(self.gw+100000,self.gh+100000)
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        # self.gview=QGraphicsView(self.scrollAreaWidgetContents)
        # self.gview.setGeometry()
        self.lal_main = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        # self.lal_main.setGeometry(QtCore.QRect(0, 0,1000,1000))
        self.lal_main.setObjectName("lal_main")

        self.hud_lt = QtWidgets.QLabel(self)
        # self.hud_lt.setWindowOpacity(0.5)
        self.hud_lt.move(10,-10)
        self.hud_lt.resize(900,200)
        self.hud_lt.setText("hello")
        self.hud_lt.setStyleSheet('color:rgb(255,0,0,255)')
        self.hud_lt.setVisible(False)

        self.scrollArea.verticalScrollBar()
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.horizontalLayout_2.addWidget(self.scrollArea)

        self.verticalLayout.addLayout(self.horizontalLayout_2)
        self.verticalLayout.setContentsMargins(0,0,0,0)

        self.horizontalLayout_3.addLayout(self.verticalLayout)
        self.line = QtWidgets.QFrame(self.centralWidget)
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.horizontalLayout_3.addWidget(self.line)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setSpacing(6)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.lal_view = QtWidgets.QLabel(self.centralWidget)
        self.lal_view.setObjectName("lal_view")
        self.lal_view.setScaledContents(True)
        self.lal_view.setText("this widget will displap a small window for the select area")
        self.verticalLayout_2.addWidget(self.lal_view)
        self.te_info = QtWidgets.QTextEdit(self.centralWidget)
        self.te_info.setObjectName("lal_info")
        self.verticalLayout_2.addWidget(self.te_info)
        self.horizontalLayout_3.addLayout(self.verticalLayout_2)
        MainWindow.setCentralWidget(self.centralWidget)

        # self.menuBar = QtWidgets.QMenuBar(MainWindow)
        # self.menuBar.setGeometry(QtCore.QRect(0, 0, 1126, 23))
        # self.menuBar.setObjectName("menuBar")
        # self.menufile = QtWidgets.QMenu(self.menuBar)
        # self.menufile.setObjectName("menufile")
        # self.menupdf = QtWidgets.QMenu(self.menuBar)
        # self.menupdf.setObjectName("menupdf")
        # MainWindow.setMenuBar(self.menuBar)

        self.statusBar = QtWidgets.QStatusBar(MainWindow)
        self.statusBar.setObjectName("statusBar")
        MainWindow.setStatusBar(self.statusBar)

        # self.actionloadpdf = QtWidgets.QAction(MainWindow)
        # self.actionloadpdf.setObjectName("actionloadpdf")
        # self.actionconvert = QtWidgets.QAction(MainWindow)
        # self.actionconvert.setObjectName("actionconvert")
        # self.actionnextpage = QtWidgets.QAction(MainWindow)
        # self.actionnextpage.setObjectName("actionnextpage")
        #
        # self.menufile.addAction(self.actionloadpdf)
        # self.menufile.addAction(self.actionconvert)
        # self.menupdf.addAction(self.actionnextpage)
        # self.menuBar.addAction(self.menufile.menuAction())
        # self.menuBar.addAction(self.menupdf.menuAction())



        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.setWindowTitle("PDF2WORD made by mengtianwxs")
        self.setWindowIcon(QIcon('p2mtw.ico'))



        self.lal_view.setFixedSize(500, 400)
        self.te_info.setFixedSize(500, self.mh - 320)


        # self.actionloadpdf.triggered.connect(self.method_loadpdf)
        # self.actionconvert.triggered.connect(self.method_convert)
        # self.actionnextpage.triggered.connect(self.method_nextpage)

        self.btn_loadpdf.clicked.connect(self.method_loadpdf)
        self.btn_convert.clicked.connect(self.method_convert)
        self.btn_nextpage.clicked.connect(self.method_nextpage)
        self.btn_prepage.clicked.connect(self.method_prepage)

        self.btn_zoomin.clicked.connect(self.method_zoominx)
        self.btn_zoomout.clicked.connect(self.method_zoomoutx)
        self.btn_rotationLeft.clicked.connect(self.method_rotationleft)
        self.btn_rotationright.clicked.connect(self.method_rotationright)

        self.btn_zoomn.clicked.connect(self.method_zoomn)

        self.btn_loadimg.clicked.connect(self.method_loadimg)

        # self.le_data.setFocus()
        self.le_data.returnPressed.connect(self.method_ledatare)
        # self.te_info.textChanged.connect(self.method_infochanged)

        # self.lal_info.textChanged.connect(self.method_infochanged)
        self.statusBar.showMessage("redlay to load  file, please push loadpdf button (CTRL + X) Or push loadIMG button (CTRL + D) ")
        self.lal_view.setText("display the  Select area Image ")
        self.lal_main.setText("To display the load file")




    def method_loadimg(self):
        self.finame, self.filtertype = QFileDialog.getOpenFileName(self, "select file", "./", "(*.png *.jpg)")
        if (self.finame != ""):
            self.pixmap=QPixmap()
            self.pixmap.load(self.finame)
            print(self.pixmap.width(),self.pixmap.height())
            if(self.pixmap.width()<10 or self.pixmap.height()<10):
                self.statusBar.showMessage("this image is too small to display, please select anthor ...")
            else:
                self.lal_main.setPixmap(self.pixmap)
                self.lal_main.resize(self.pixmap.width(),self.pixmap.height())
                self.nsizew=self.pixmap.width()
                self.nsizeh=self.pixmap.height()
                self.openmethod="mimg"
                self.npixmap=self.pixmap
                self.statusBar.showMessage("the software has loaded image file ,please drag your mouse to select area ...")

        else:
            self.statusBar.showMessage("has no image file selected , please select one image file ...")

    def method_zoominx(self):
        if (os.path.exists(self.finame)):
            ppimg=PImage.open(self.finame)
            w,h=ppimg.size
            cpimg=ppimg.copy()
            newimg=cpimg.resize((int(w*1.5),int(h*1.5)))
            nw,nh=newimg.size
            newimg.save(self.img_name_zoomin_saved)
            self.pimg=QPixmap()
            self.pimg.load(self.img_name_zoomin_saved)
            self.lal_main.setPixmap(self.pimg)
            self.lal_main.resize(self.pimg.width(),self.pimg.height())

    def method_zoomoutx(self):
        if (os.path.exists(self.finame)):
            ppimg = PImage.open(self.finame)
            w, h = ppimg.size
            cpimg = ppimg.copy()
            newimg = cpimg.resize((int(w /1.5), int(h / 1.5)))
            nw, nh = newimg.size
            newimg.save(self.img_name_zoomout_saved)
            self.pimg = QPixmap()
            self.pimg.load(self.img_name_zoomout_saved)
            self.lal_main.setPixmap(self.pimg)
            self.lal_main.resize(self.pimg.width(), self.pimg.height())

    def method_rotationleft(self):
        if (os.path.exists(self.finame)):
            ppima=PImage.open(self.finame)
            w,h=ppima.size
            print(w,h)
            lrimg=ppima.rotate(90)
            # nlrimg=lrimg.resize((h*2,w*2))
            # print(nlrimg.size)
            lrimg.save(self.img_name_rotate_left)
            self.plimg=QPixmap()
            self.plimg.load(self.img_name_rotate_left)
            self.lal_main.setPixmap(self.plimg)
            self.lal_main.resize(lrimg.width,lrimg.height)

    def method_rotationright(self):
        if (os.path.exists(self.finame)):
            ppima = PImage.open(self.finame)
            lrimg = ppima.rotate(-90)
            lrimg.save(self.img_name_rotate_right)
            self.primg = QPixmap()
            self.primg.load(self.img_name_rotate_right)
            self.lal_main.setPixmap(self.primg)
            self.lal_main.resize(self.primg.width(), self.primg.height())

    def method_zoomn(self):
        if(self.nsizeh !=0 and self.nsizew !=0):
            self.lal_main.setPixmap(self.npixmap)
            self.lal_main.resize(self.npixmap.width(),self.npixmap.height())

    def method_infochanged(self):
        self.listdata.clear()
        str = self.te_info.toPlainText()
        self.ls_content = str.split('\n')
        # print(str)
        # print('infochange ',self.ls_content)
        self.listdata=self.ls_content
        self.statusBar.showMessage("has refresh and changed data")


    def method_startcad(self):
        self.statusBar.showMessage("please write it is start autocad application and init ...")
        self.mtdwg = mtp2autocad.p2cad()
        time.sleep(3)


    def method_ledatare(self):
        self.letxt=self.le_data.text()
        if(self.fname!=""):
            if(self.letxt=="end"):
                self.img = pdf2image._run_convert(self.fname, self.pageNum-1)
                if (os.path.exists(self.img_name_convert_saved)):
                    self.lal_main.setPixmap(QPixmap(self.img_name_convert_saved))
                    self.inx = self.pageNum-1
            elif(self.letxt=="start" or self.letxt=="0"):
                self.img = pdf2image._run_convert(self.fname, 0)
                if (os.path.exists(self.img_name_convert_saved)):
                    self.lal_main.setPixmap(QPixmap(self.img_name_convert_saved))
                    self.inx = 0

        if (self.fname != "" and self.letxt.isdigit()):
            # print("is digit")
            self.inxle=int(self.letxt)
            if(self.inxle>=0 and self.inxle <self.pageNum):
                self.img = pdf2image._run_convert(self.fname, self.inxle)
                if (os.path.exists(self.img_name_convert_saved)):
                    self.lal_main.setPixmap(QPixmap(self.img_name_convert_saved))
                    self.inx=self.inxle


            else:
                self.statusBar.showMessage("index is out ...")



        else:
            self.statusBar.showMessage("prepage, please load pdf file first")
        self.method_displayHUD()
        self.lal_main.setFocus()

    def method_prepage(self):
        if(self.fname!=""):
            if (self.pageNum == 1):
                self.statusBar.showMessage("total 1 page")
            elif self.pageNum > 1 :
                # if(self.inxle!=0):
                #     self.inx=self.inxle
                self.inx = self.inx - 1

                self.statusBar.showMessage("page: "+str(self.pageNum)+'/'+str(self.inx))
                print(self.inx)
                if self.inx < self.pageNum and self.inx >=0:
                    self.img = pdf2image._run_convert(self.fname, self.inx)
                    if (os.path.exists(self.img_name_convert_saved)):
                        self.lal_main.setPixmap(QPixmap(self.img_name_convert_saved))
                else:
                    self.inx=0
                    self.statusBar.showMessage("index out ")
            self.method_displayHUD()
        else:
            self.statusBar.showMessage("prepage, please load pdf file first")

    def method_nextpage(self):
        if(self.fname != ""):
            if(self.pageNum==1):
                self.statusBar.showMessage("total 1 page")
            elif self.pageNum > 1:
                # print("this is page 2")
                self.inx = self.inx + 1

                self.statusBar.showMessage("page: "+str(self.pageNum)+'/'+str(self.inx))
                if self.inx <self.pageNum:
                    self.img = pdf2image._run_convert(self.fname, self.inx)
                    if (os.path.exists(self.img_name_convert_saved)):
                        self.lal_main.setPixmap(QPixmap(self.img_name_convert_saved))
                else:
                    self.inx=0
                    self.statusBar.showMessage("index out ")
            self.method_displayHUD()
        else:
            self.statusBar.showMessage("nextpage, please load pdf file first")

    def method_convert(self):
        if(self.fname != "" or self.finame !=""):
            if(self.aabb==True):
                self.content = self.bdocr.invoke(self.img_name_orc_invoke)
                for i in range(0, len(self.content)):
                    # print(str(self.content[i]['words']))
                    self.te_info.setPlainText(self.te_info.toPlainText() + str(self.content[i]['words']) + '\n')
                    self.listdata.append(str(self.content[i]['words']))
                    cursor=self.te_info.textCursor()
                    cursor.movePosition(QtGui.QTextCursor.End)
                    self.te_info.setTextCursor(cursor)

                if (os.path.exists(self.img_name_orc_invoke)):
                    os.remove(self.img_name_orc_invoke)
                    # print("del img")
                    self.aabb = False
                    self.statusBar.showMessage("del a img")
                    time.sleep(1)
                    self.statusBar.showMessage("complete convert words, and you can press CTRL + S to save the data with a excel file")
                else:
                    self.statusBar.showMessage("there is no pic a img")
        else:
            self.statusBar.showMessage("convert ,please load pdf file first ...")

    def method_loadpdf(self):
        self.fname, self.filtertype = QFileDialog.getOpenFileName(self, "select file", "d://", "*.pdf")
        if(self.fname!=""):
            self.img = pdf2image._run_convert(self.fname, self.inx)
            # print(self.img.width,self.img.height)
            self.lal_main.resize(self.img.width, self.img.height)
            self.nsizeh=self.img.width
            self.nsizew=self.img.height
            self.openmethod="mpdf"
            self.ffname=self.fname

            if (os.path.exists(self.img_name_convert_saved)):
                self.pp=QPixmap()
                self.pp.load(self.img_name_convert_saved)
                self.lal_main.setPixmap(self.pp)
                self.npixmap=self.pp
                self.finame=self.img_name_convert_saved
            self.pageNum = pdf2image.getPDF(self.fname).getNumPages()
            self.statusBar.showMessage("the software has loaded pdf file ,please drag your mouse to select area ...")
            # self.le_data.setFocus()

        else:
            self.statusBar.showMessage("has no pdf file selected , please select one pdf file ...")

    def method_save(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "0"
        for i in range(0, len(self.listdata)):
            ws['A' + str(i + 1)] = self.listdata[i]
        ti = time.time()
        strs = self.fname
        ix = strs.rfind('/')
        newstr = strs[0:ix + 1]
        wname = newstr + 'p' + str(ti)[:2] + '.xlsx'
        wb.save(wname)
        self.statusBar.showMessage("has saved the file in  " + wname)

    def method_dwgToSave(self):
        if(len(self.listdata)>0):
            self.mtdwg.loadData(self.listdata)
            dwginfo=self.mtdwg.pasreAndDrawToAutoCad()
            self.statusBar.showMessage(dwginfo)

    def method_insertat(self):
        self.te_info.append("@\n")
        # pass

    def keyReleaseEvent(self, e):
        pass
        # if e.key()==QtCore.Qt.Key_Space:
        #     self.vdd = self.posy2 - self.posy1
        #     self.hdd = self.posx2 - self.posx1
        #     print(self.posy2,self.posy1,self.vdd)
        #     print(self.posx2,self.posx1,self.hdd)
        #
        #     if (self.vdd > 0):
        #         self.scrollArea.verticalScrollBar().setValue(self.vd + self.vdd)
        #     else:
        #         self.scrollArea.verticalScrollBar().setValue(self.vd - self.vdd)
        #
        #     if(self.hdd>0):
        #         self.scrollArea.horizontalScrollBar().setValue(self.hd+self.hdd)
        #     else:
        #         self.scrollArea.horizontalScrollBar().setValue(self.hd-self.hdd)

    def method_displayHUD(self):
        fn = ''  # file name
        pn = 0  # page num
        ind = 0  # index page

        if (self.pageNum != 0):
            pn = self.pageNum
            ind = self.inx
        self.hud_lt.setText("file name : " + self.ffname + '\n' + 'total page : ' + str(pn) + '\n' + 'index page : ' + str(ind))

    def keyPressEvent(self,e):

        # if e.key() == QtCore.Qt.Key_Space:
        #     print("space", self.scrollArea.verticalScrollBar().value())
        #     # self.scrollArea.verticalScrollBar().setValue(500)
        #     self.vd = self.scrollArea.verticalScrollBar().value()
        #     self.hd = self.scrollArea.horizontalScrollBar().value()

        if e.key()==QtCore.Qt.Key_W:
            self.le_data.clear()
            self.le_data.setFocus()

        if e.key()==QtCore.Qt.Key_R:
            if QApplication.keyboardModifiers()==QtCore.Qt.ControlModifier:
                self.method_infochanged()


        if e.key()==QtCore.Qt.Key_S:
            if QApplication.keyboardModifiers()==QtCore.Qt.ControlModifier:
                self.method_save()

        # load img
        if e.key()==QtCore.Qt.Key_D:
            if QApplication.keyboardModifiers()==QtCore.Qt.ControlModifier:
                self.method_loadimg()
                # load img

        if e.key()==QtCore.Qt.Key_Q:
            if QApplication.keyboardModifiers() == QtCore.Qt.ControlModifier:
                print("start cad")
                self.method_startcad()


                # load img
        if e.key() == QtCore.Qt.Key_B:
            if QApplication.keyboardModifiers() == QtCore.Qt.ControlModifier:
                self.method_insertat()
                # load img

        if e.key() == QtCore.Qt.Key_E:
            if QApplication.keyboardModifiers() == QtCore.Qt.ControlModifier:
                self.lal_main.setFocus()
                # load img
        if e.key()==QtCore.Qt.Key_T:
            if QApplication.keyboardModifiers() == QtCore.Qt.ControlModifier:

                if(self.bhdu):
                    self.hud_lt.setVisible(True)
                    self.bhdu = False
                else:
                    self.hud_lt.setVisible(False)
                    self.bhdu=True
                self.method_displayHUD()

        if e.key() == QtCore.Qt.Key_W:
            if QApplication.keyboardModifiers() == QtCore.Qt.ControlModifier:
                self.method_dwgToSave()

        if e.key()==QtCore.Qt.Key_X:
            if QApplication.keyboardModifiers()==QtCore.Qt.AltModifier:
                self.setGeometry(0,0,self.gw,self.gh)

            if QApplication.keyboardModifiers()==QtCore.Qt.ControlModifier:
                self.method_loadpdf()

        if e.key()==QtCore.Qt.Key_C:
            if QApplication.keyboardModifiers()==QtCore.Qt.AltModifier:
                self.setGeometry(self.mposx,self.mposy,self.mw,self.mh)

            if QApplication.keyboardModifiers() == QtCore.Qt.ControlModifier:
                self.method_convert()

        if e.key()==QtCore.Qt.Key_X:
            if QApplication.keyboardModifiers()==QtCore.Qt.AltModifier:
                self.setGeometry(0,0,self.gw,self.gh)

        if e.key() == QtCore.Qt.Key_N:
            self.lal_main.setFocus()
            self.method_nextpage()

        if e.key() == QtCore.Qt.Key_P:
            self.lal_main.setFocus()

            self.method_prepage()


        if e.key()==QtCore.Qt.Key_J:
            self.lal_main.setFocus()

            self.dist = self.dist - 100
            # print(self.dist)

            v = self.scrollArea.verticalScrollBar().value()
            if (v >= 0):
                self.scrollArea.verticalScrollBar().setValue(self.dist)
            else:
                self.dist = 0

        if e.key()==QtCore.Qt.Key_K:
            self.lal_main.setFocus()

            self.dist = self.dist +100
            # print(self.dist)

            v = self.scrollArea.verticalScrollBar().value()
            if (v >= 0):
                self.scrollArea.verticalScrollBar().setValue(self.dist)
            else:
                self.dist = 0

        if e.key() == QtCore.Qt.Key_H:
            self.lal_main.setFocus()

            self.dist=self.dist-100
            print(self.dist)

            v=self.scrollArea.horizontalScrollBar().value()
            if(v>=0):
                self.scrollArea.horizontalScrollBar().setValue(self.dist)
            else:
                self.dist=0

        if e.key() == QtCore.Qt.Key_L:
            self.lal_main.setFocus()

            self.dist=self.dist+100

            v = self.scrollArea.horizontalScrollBar().value()
            if (v >=0):
                self.scrollArea.horizontalScrollBar().setValue(self.dist)
            else:
                self.dist=0

    def mousePressEvent(self,e):
        if e.button()==QtCore.Qt.LeftButton:
            self.posx1=pyautogui.position()[0]
            self.posy1=pyautogui.position()[1]
            self.count=self.count+1

            self.sx1=e.pos().x()
            print(self.sx1)
            self.vd = self.scrollArea.verticalScrollBar().value()
            self.hd = self.scrollArea.horizontalScrollBar().value()

            # self.statusBar.showMessage('mouse has press,and the start postion is ('+str(self.posx1)+','+str(self.posy1)+')')

    def mouseReleaseEvent(self, e):

        # if e.button()==QtCore.Qt.MiddleButton:
        #     pass


        if e.button()==QtCore.Qt.LeftButton:
            self.posx2=pyautogui.position()[0]
            self.posy2=pyautogui.position()[1]
            self.relWidth=self.posx2-self.posx1
            self.relHeight=self.posy2-self.posy1
            self.relarea=self.relWidth*self.relHeight
            self.statusBar.showMessage('mouse has press,and the start postion is ('+str(self.posx1)+','+str(self.posy1)+')'+
                                       ',and the end postion is ('+str(self.posx2)+','+str(self.posy2)+')'+' and area (W,H)=( '
                                       +str(self.relWidth)+','+str(self.relHeight)+' )'+'='+str(self.relarea))

            if(self.relarea>100):
                # print(self.posx1, self.posy1, self.count)
                # print(self.posx2, self.posy2)
                # print(self.relWidth, self.relHeight)
                im = pyautogui.screenshot(region=(self.posx1, self.posy1, self.relWidth, self.relHeight))
                im.save(self.img_name_orc_invoke)
                self.lal_view.setPixmap(QPixmap(self.img_name_orc_invoke))
                self.aabb=True
                # print("save img success")
                time.sleep(2)
                self.statusBar.showMessage("save a img success, please press convert button or press CTRL + C")
            else:
                # print("this is error")
                self.statusBar.showMessage("this area is too small to get word")

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.btn_loadpdf.setText(_translate("MainWindow", "loadpdf"))
        self.btn_convert.setText(_translate("MainWindow", "convert"))
        self.btn_nextpage.setText(_translate("MainWindow", "nextpage"))
        self.btn_prepage.setText(_translate("MainWindow","prepage"))
        self.btn_rotationright.setText(_translate("MainWindow","rotationR"))
        self.btn_rotationLeft.setText(_translate("MainWindow","rotationL"))
        self.btn_zoomin.setText(_translate("MainWindow", "zoomin"))
        self.btn_zoomout.setText(_translate("MainWindow","zoomout"))
        self.btn_loadimg.setText(_translate("MainWindow","loadIMG"))
        self.btn_zoomn.setText(_translate("MainWindow","normal"))



        self.lal_main.setText(_translate("MainWindow", "TextLabel"))
        self.lal_view.setText(_translate("MainWindow", "TextLabel"))


        # self.menufile.setTitle(_translate("MainWindow", "run"))
        # self.menupdf.setTitle(_translate("MainWindow", "pdf"))
        # self.actionloadpdf.setText(_translate("MainWindow", "loadpdf"))
        # self.actionconvert.setText(_translate("MainWindow", "convert"))
        # self.actionnextpage.setText(_translate("MainWindow", "nextpage"))

if __name__=="__main__":
    app=QApplication(sys.argv)
    mw=Ui_MainWindow()
    mw.show()
    sys.exit(app.exec_())