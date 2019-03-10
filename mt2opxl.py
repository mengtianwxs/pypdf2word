from openpyxl import Workbook,load_workbook
import openpyxl
from openpyxl.styles import NamedStyle, Side, Border


# wb=Workbook()
# sheet=wb.active
# listdata=['@ggd1','fjxg','@ggd2','fix2']
# sheet.merge_cells('A1:I1')
# sheet['A1'].value='ggd1'
# sheet.append(['序号','元件名称','型号规格','单位','数量','单   价','总   价','生产厂家','备注'])
# highlight=NamedStyle(name='highlight')
# bd=Side(style='thick',color='000000')
# highlight.border=Border(left=bd,right=bd,top=bd,bottom=bd)

# sheet['A2':'I2'].style=highlight


# wb.save("./aabb.xlsx")

class p2openxl:
    def __init__(self):
        self.data=[]

    def loadDataAndParse(self,nxl):

        wb=openpyxl.load_workbook(nxl)
        st=wb.active
        mx_row=st.max_row
        print(mx_row,type(mx_row))
        for i in range(0,mx_row):
            # print(st['A'+str(i+1)].value)
            self.data.append(st['A'+str(i+1)].value)

        return self.data




    def saveData(self,lstdata,namexls):
        wb = Workbook()
        ws = wb.active
        for i in range(0, len(lstdata)):
            ws['A' + str(i + 1)] = lstdata[i]

        wb.save(namexls)




